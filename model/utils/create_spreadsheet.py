import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from config import EXPORT_FOLDER

def create_spreadsheet(filename, timestamp_str):

    # Read the sim results
    df = pd.read_csv(filename)

    # df = df.rename(columns={'price': 'Odds', 'points': 'O/U', 'player_name': 'Name', 'player_team': 'Team', 'defensive_matchup': 'Opp', 'sports_books': 'Sportsbook', 'edge': 'Edge'}, inplace=True)

    # Filter records where edge => 0.12 (12%)
    filtered_df = df[df['edge'] >= 0.12]

    # Keep the record with the highest odds for each player_name
    df_unique = filtered_df.sort_values('price', ascending=False).drop_duplicates('player_name', keep='first')

    # Remove the specified columns
    columns_to_remove = ['id', 'prop_type', 'nba_api_player_id', 'implied_odds', 'p(over)', 'p(under)']
    df_cleaned = df_unique.drop(columns=columns_to_remove)

    # Sort by edge in descending order for better readability
    sorted_df = df_cleaned.sort_values('edge', ascending=False)

    # Create a new workbook
    wb = create_formatted_workbook(sorted_df)

    wb.save(EXPORT_FOLDER + f'/worksheets/{timestamp_str}.xlsx')

# Function to create formatted workbook
def create_formatted_workbook(dataframe):
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "NBA 3 Point Model"
    
    # Write the dataframe to the worksheet
    for r in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(r)

    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Find the edge column
    edge_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'edge':
            edge_col_idx = idx
            break

    # Define color fills
    dark_yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dark yellow for 12%-15
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green for 16-22%
    dark_green_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")   # Dark green for > 22%

    # Apply conditional formatting to edge column
    if edge_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=edge_col_idx)
            edge_value = cell.value
            
            if edge_value is not None:
                if edge_value > 0.22:  # Dark green for > 22%
                    cell.fill = dark_green_fill
                elif 0.15 < edge_value <= 0.22:  # Light green for > 15% and <= 22%
                    cell.fill = light_green_fill
                elif 0.12 <= edge_value <= 0.15:  # Dark yellow for 12% to 15%
                    cell.fill = dark_yellow_fill

    # Adjust column widths
    column_widths = {
        'A': 18,  # player_name
        'B': 12,  # player_team
        'C': 18,  # defensive_matchup
        'D': 12,  # sports_book
        'E': 8,   # name
        'F': 8,   # odds
        'G': 8,   # o/u
        'H': 10,  # edge
        'I': 10,  # ev
        'J': 16   # suggested_kelly
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    return wb